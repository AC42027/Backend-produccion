from django.contrib import admin, messages
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.urls import path

from import_export import resources
from import_export.admin import ImportExportModelAdmin
from import_export.fields import Field
from import_export.widgets import ForeignKeyWidget
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import (
    Division, Area, Zona, Equipo, Inspeccion, UbicacionFisica,
    Categoria, InspeccionTecnico, PreguntaTecnica, Owner
)
from . import sap_assets

# ---------- EQUIPOS: carga masiva ----------
# Catálogos usados por la plantilla Excel y el Resource.
# (columna, modelo, campo_del_catálogo, nombre_de_hoja)
PLANTILLA_CATALOGOS = (
    ("division", Division, "nombre", "Divisiones"),
    ("area", Area, "nombre", "Areas"),
    ("zona", Zona, "nombre", "Zonas"),
    ("owner", Owner, "nombre", "Owners"),
    ("categoria", Categoria, "nombre", "Categorias"),
    ("ubicacion", UbicacionFisica, "descripcion", "Ubicaciones"),
)

PLANTILLA_HOJA = "Equipos"
PLANTILLA_MAX_FILAS = 1000


class EquipoResource(resources.ModelResource):
    division = Field(
        column_name="division", attribute="division",
        widget=ForeignKeyWidget(Division, "nombre"),
    )
    area = Field(
        column_name="area", attribute="area",
        widget=ForeignKeyWidget(Area, "nombre"),
    )
    zona = Field(
        column_name="zona", attribute="zona",
        widget=ForeignKeyWidget(Zona, "nombre"),
    )
    owner = Field(
        column_name="owner", attribute="owner",
        widget=ForeignKeyWidget(Owner, "nombre"),
    )
    categoria = Field(
        column_name="categoria", attribute="categoria",
        widget=ForeignKeyWidget(Categoria, "nombre"),
    )
    ubicacion = Field(
        column_name="ubicacion", attribute="ubicacion",
        widget=ForeignKeyWidget(UbicacionFisica, "descripcion"),
    )

    class Meta:
        model = Equipo
        fields = (
            "nombre",
            "division", "area", "zona",
            "owner", "categoria", "ubicacion",
            "sap_equnr", "sap_equnr_desc", "sap_tplnr", "sap_tplnr_desc",
        )
        # El nombre identifica al equipo: existente -> se actualiza, no se duplica
        import_id_fields = ("nombre",)
        skip_unchanged = True
        report_skipped = True
        clean_model_instances = True

    def before_save_instance(self, instance, row, **kwargs):
        """Auto-asigna datos SAP según el nombre si la fila no los trae."""
        if not instance.sap_equnr and not instance.sap_tplnr and instance.nombre:
            try:
                auto = sap_assets.auto_match(instance.nombre)
            except Exception:
                auto = None  # la API nunca debe impedir importar
            if auto:
                instance.sap_equnr = auto['equnr']
                instance.sap_equnr_desc = auto['equnr_desc']
                instance.sap_tplnr = auto['tplnr']
                instance.sap_tplnr_desc = auto['tplnr_desc']
        super().before_save_instance(instance, row, **kwargs)


# ---------- EQUIPOS ----------
@admin.register(Equipo)
class EquipoAdmin(ImportExportModelAdmin):
    resource_classes = [EquipoResource]
    change_list_template = "admin/inspeccion/equipo/change_list.html"
    list_display = (
        "id", "nombre",
        "division_nombre", "area_nombre", "zona_nombre",
        "owner_nombre", "categoria_nombre", "ubicacion_descripcion",
        "sap_tplnr",
    )
    list_filter = (
        ("division", admin.RelatedOnlyFieldListFilter),
        ("area", admin.RelatedOnlyFieldListFilter),
        ("zona", admin.RelatedOnlyFieldListFilter),
        ("owner", admin.RelatedOnlyFieldListFilter),
        ("categoria", admin.RelatedOnlyFieldListFilter),
    )
    search_fields = (
        "nombre",
        "division__nombre", "area__nombre", "zona__nombre",
        "owner__nombre", "categoria__nombre",
        "ubicacion__descripcion",
        "sap_equnr", "sap_tplnr",
    )
    list_select_related = ("division", "area", "zona", "owner", "categoria", "ubicacion")
    ordering = ("division__nombre", "area__nombre", "zona__nombre", "nombre")
    autocomplete_fields = ("division", "area", "zona", "owner", "categoria", "ubicacion")

    class Media:
        js = ("inspeccion/admin_sap_search.js",)

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "descargar-plantilla/",
                self.admin_site.admin_view(self.descargar_plantilla),
                name="inspeccion_equipo_plantilla",
            ),
        ]
        return custom + urls

    def descargar_plantilla(self, request):
        """Genera un .xlsx con dropdowns reales (validación de datos) para
        las columnas division/area/zona/owner/categoria/ubicacion."""
        if not self.has_add_permission(request):
            raise PermissionDenied

        wb = Workbook()
        ws = wb.active
        ws.title = PLANTILLA_HOJA

        headers = ["nombre"] + [campo for campo, *_ in PLANTILLA_CATALOGOS]
        ws.append(headers)
        for celda in ws[1]:
            celda.font = Font(bold=True)
        ws.freeze_panes = "A2"
        anchos = {"nombre": 35}
        for campo, *_ in PLANTILLA_CATALOGOS:
            anchos[campo] = 25
        for idx, ancho in enumerate(anchos.values(), start=1):
            ws.column_dimensions[get_column_letter(idx)].width = ancho

        for col_idx, (campo, modelo, campo_catalogo, hoja) in enumerate(
            PLANTILLA_CATALOGOS, start=2
        ):
            cs = wb.create_sheet(hoja)
            valores = list(
                modelo.objects.order_by(campo_catalogo)
                .values_list(campo_catalogo, flat=True)
            )
            titulo = cs.cell(row=1, column=1, value=campo)
            titulo.font = Font(bold=True)
            cs.column_dimensions["A"].width = max(25, *(len(str(v)) + 4 for v in valores)) if valores else 25
            for fila, valor in enumerate(valores, start=2):
                cs.cell(row=fila, column=1, value=valor)

            if not valores:
                continue
            dv = DataValidation(
                type="list",
                formula1=f"'{hoja}'!$A$2:$A${len(valores) + 1}",
                allow_blank=True,
                errorStyle="warning",
                errorTitle="Valor no encontrado",
                error="El valor no está en el catálogo actual; "
                      "revise la hoja '%s' o créelo en el admin." % hoja,
            )
            letra = get_column_letter(col_idx)
            dv.add(f"{letra}2:{letra}{PLANTILLA_MAX_FILAS}")
            ws.add_data_validation(dv)

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = (
            'attachment; filename="plantilla_equipos.xlsx"'
        )
        wb.save(response)
        return response

    def save_model(self, request, obj, form, change):
        """Auto-asigna datos SAP según el nombre si los campos están vacíos."""
        auto = None
        sin_datos = not obj.sap_equnr and not obj.sap_tplnr
        if sin_datos and obj.nombre:
            try:
                auto = sap_assets.auto_match(obj.nombre)
            except Exception:
                auto = None  # la API nunca debe impedir guardar
            if auto:
                obj.sap_equnr = auto['equnr']
                obj.sap_equnr_desc = auto['equnr_desc']
                obj.sap_tplnr = auto['tplnr']
                obj.sap_tplnr_desc = auto['tplnr_desc']
        super().save_model(request, obj, form, change)
        if auto:
            destino = auto['tplnr'] or auto['equnr']
            messages.success(
                request,
                f'Ubicación técnica asignada automáticamente desde SAP: {destino}'
            )
        elif sin_datos:
            messages.warning(
                request,
                'No se encontró este equipo en SAP; '
                'asígnelo manualmente con el buscador "Buscar en SAP".'
            )

    # columnas legibles
    def division_nombre(self, obj): return getattr(obj.division, "nombre", "")
    division_nombre.short_description = "División"
    division_nombre.admin_order_field = "division__nombre"

    def area_nombre(self, obj): return getattr(obj.area, "nombre", "")
    area_nombre.short_description = "Área"
    area_nombre.admin_order_field = "area__nombre"

    def zona_nombre(self, obj): return getattr(obj.zona, "nombre", "")
    zona_nombre.short_description = "Zona"
    zona_nombre.admin_order_field = "zona__nombre"

    def owner_nombre(self, obj): return getattr(obj.owner, "nombre", "")
    owner_nombre.short_description = "Owner"
    owner_nombre.admin_order_field = "owner__nombre"

    def categoria_nombre(self, obj): return getattr(obj.categoria, "nombre", "")
    categoria_nombre.short_description = "Categoría"
    categoria_nombre.admin_order_field = "categoria__nombre"

    def ubicacion_descripcion(self, obj): return getattr(obj.ubicacion, "descripcion", "")
    ubicacion_descripcion.short_description = "Ubicación física"
    ubicacion_descripcion.admin_order_field = "ubicacion__descripcion"


# ---------- RESTO (opcional, más útil con filtros/búsqueda) ----------
@admin.register(Division)
class DivisionAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre")
    search_fields = ("nombre",)
    ordering = ("nombre",)

@admin.register(Area)
class AreaAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre", "division")
    list_filter = (("division", admin.RelatedOnlyFieldListFilter),)
    search_fields = ("nombre", "division__nombre")
    list_select_related = ("division",)
    ordering = ("division__nombre", "nombre")

@admin.register(Zona)
class ZonaAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre", "area")
    list_filter = (("area", admin.RelatedOnlyFieldListFilter),)
    search_fields = ("nombre", "area__nombre")
    list_select_related = ("area",)
    ordering = ("area__nombre", "nombre")

@admin.register(UbicacionFisica)
class UbicacionFisicaAdmin(admin.ModelAdmin):
    list_display = ("id", "descripcion")
    search_fields = ("descripcion",)
    ordering = ("descripcion",)

@admin.register(Categoria)
class CategoriaAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre")
    search_fields = ("nombre",)
    ordering = ("nombre",)

@admin.register(Owner)
class OwnerAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre")
    search_fields = ("nombre",)
    ordering = ("nombre",)

@admin.register(Inspeccion)
class InspeccionAdmin(admin.ModelAdmin):
    list_display = ('fecha', 'equipo', 'sap_equnr', 'sap_puesto_trabajo', 'owner', 'sap_nr_numero', 'sap_nr_status')
    list_filter  = ("fecha", ("equipo", admin.RelatedOnlyFieldListFilter), "sap_nr_status")
    search_fields = ("equipo__nombre", "sap_equnr", "sap_puesto_trabajo", "sap_nr_numero")
    readonly_fields = ('sap_nr_numero', 'sap_nr_status')
    fields = (
        'fecha', 'hora_inicio', 'hora_fin',
        'division', 'area', 'zona', 'equipo',
        'sap_equnr', 'sap_equnr_desc', 'sap_tplnr', 'sap_puesto_trabajo',
        'observaciones', 'comentario_hallazgo',
        'owner',
        'sap_nr_numero', 'sap_nr_status',   # campos de resultado SAP (solo lectura)
    )

@admin.register(InspeccionTecnico)
class InspeccionTecnicoAdmin(admin.ModelAdmin):
    list_display = ("id", "inspeccion", "descripcion", "estado")
    list_filter = ("estado",)

@admin.register(PreguntaTecnica)
class PreguntaTecnicaAdmin(admin.ModelAdmin):
    list_display = ("id", "descripcion", "categoria")
    list_filter = (("categoria", admin.RelatedOnlyFieldListFilter),)
    search_fields = ("descripcion", "categoria__nombre")
